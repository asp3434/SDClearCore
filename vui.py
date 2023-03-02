import tkinter
import customtkinter
import tkinter.messagebox
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import (FigureCanvasTkAgg, NavigationToolbar2Tk)
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment
import time
import re
import serial



##########################
# Evan Peck
# Capstone Project
# 2/11/23
# Group 6, Skittlz Dynasty
# GUI for control of Generant Compaction Station
##########################

###### NOTES ########
# Save values as variables
# Add AutoLiv and UofU images
# Get feedback on color pallet?
# Add opening window w/ images, manual/automatic
# Find frequency of rotation direction change
# Error message for invalid filename
# Fix errors in filesave if not all subsystems are active

customtkinter.set_appearance_mode("dark")
customtkinter.set_default_color_theme("blue")

freqBounds1 = [50, 100]
freqBounds06 = [50, 125]
freqBounds04 = [50, 150]
freqBoundsCustom = [50, 150]
freqBounds = [freqBounds1[0], freqBounds1[1]]

forceBounds1 = [0, 300]
forceBounds06 = [0, 400]
forceBounds04 = [0, 500]
forceBoundsCustom = [0, 550]
forceBounds = [forceBounds1[0], forceBounds1[1]]

rotSpeedBounds1 = [0, 30]
rotSpeedBounds06 = [0, 40]
rotSpeedBounds04 = [0, 50]
rotSpeedBoundsCustom = [0, 55]
rotSpeedBounds = [rotSpeedBounds1[0], rotSpeedBounds1[1]]

timeBounds = [0, 60]

freqResult = 75
forceResult = 250
rotateResult = 9.81
timeResult = 5

clearcoreResults = None
clearcore = serial.Serial('COM3', 9600)

#Temp measurements
t = np.arange(0.0, 3.0, 0.01)
freqMeasures = np.sin(np.pi * t) + 80
forceMeasures = np.cos(np.pi * t)+150
rotationMeasures = 9.8*np.sin(np.pi * t*5)
maxFreq = max(freqMeasures)
maxForce = max(forceMeasures)
maxRotation = max(rotationMeasures)
maxTime = max(t)


class ResultsTab(customtkinter.CTkTabview):
    def __init__(self, master, **kwargs):
        super().__init__(master, **kwargs)

        global t, freqMeasures, forceMeasures, rotationMeasures

        #Create tabs
        self.add("Measured Results")
        self.add("Freq. Results")
        self.add("Force Results")
        self.add("Rot. Speed Results")

        #Insert results in tabs
        measured_freq = 'Measured frequency:               ' + str(freqResult) + 'Hz\n\n'
        measured_force = 'Measured force:                        ' + str(forceResult) + 'N\n\n'
        measured_rotation = "Measured rotation speed:       " + str(rotateResult) + "rad/s\n\n"
        measured_time = "Measured run time:                   " + str(timeResult) + "sec."
        self.measuredText = customtkinter.CTkTextbox(master=self.tab("Measured Results"))
        self.measuredText.pack(fill='both', padx=10, pady=10)
        self.measuredText.insert("0.0", measured_freq + measured_force + measured_rotation + measured_time)

        # Frequency plot
        freqFig = Figure(figsize=(8, 5))  # Frequency plot
        freqPlot = freqFig.add_subplot(111)
        # t = np.arange(0.0, 3.0, 0.01)
        # freqMeasures = np.sin(np.pi * t)+80
        #maxFreq = max(freqMeasures)
        freqPlot.axhline(y=maxFreq, color='r', linestyle='--')
        freqPlot.annotate(str(maxFreq), xy=(t[-1], maxFreq), xytext=(20,0), color='r',
                          textcoords="offset points", va='center')
        freqPlot.plot(t, freqMeasures)
        freqPlot.set(xlabel='Time [sec.]', ylabel='Freq. [Hz]')

        freqCanvas = FigureCanvasTkAgg(freqFig, master=self.tab("Freq. Results"))
        freqCanvas.draw()
        freqCanvas.get_tk_widget().pack(fill='both', padx=10, pady=10)

        #Force plot
        forceFig = Figure(figsize=(5,5))
        forcePlot = forceFig.add_subplot(111)
        #forceMeasures = np.cos(np.pi * t)+150
        #maxForce = max(forceMeasures)
        forcePlot.axhline(y=maxForce, color='r', linestyle='--')
        forcePlot.annotate(str(maxForce), xy=(t[-1], maxForce), xytext=(20, 0), color='r',
                          textcoords="offset points", va='center')
        forcePlot.plot(t, forceMeasures)
        forcePlot.set(xlabel='Time [sec.]', ylabel='Force [N]')

        forceCanvas = FigureCanvasTkAgg(forceFig, master=self.tab("Force Results"))
        forceCanvas.draw()
        forceCanvas.get_tk_widget().pack(fill='both', padx=10, pady=10)

        #Rotation plot
        #Derive data to find frequency of direction change
        rotationFig = Figure(figsize=(5, 5))
        rotationPlot = rotationFig.add_subplot(111)
        #rotationMeasures = 9.8*np.sin(np.pi * t*5)
        #maxRotation = max(rotationMeasures)
        rotationPlot.axhline(y=maxRotation, color='r', linestyle='--')
        rotationPlot.annotate(str(maxRotation), xy=(t[-1], maxForce), xytext=(20, 0), color='r',
                           textcoords="offset points", va='center')
        rotationPlot.plot(t, rotationMeasures)
        rotationPlot.set(xlabel='Time [sec.]', ylabel='Speed [rad/s]')

        rotationCanvas = FigureCanvasTkAgg(rotationFig, master=self.tab("Rot. Speed Results"))
        rotationCanvas.draw()
        rotationCanvas.get_tk_widget().pack(fill='both', padx=10, pady=10)

class ProcessResults(customtkinter.CTkToplevel):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.title("Process Results")   #Change to Filename
        self.geometry("700x500")
        # self.grid_columnconfigure(0, weight=1)
        # self.grid_rowconfigure(0, weight=1)
        # self.resultsWindowFrame = customtkinter.CTkFrame(master=self)
        # self.resultsWindowFrame.grid(row=0, column=0, padx=10, pady=10, sticky='nsew')

        self.tab_view = ResultsTab(master=self)
        self.tab_view.pack(fill='both', expand=True, padx=10, pady=10)



class App(customtkinter.CTk):
    def __init__(self):
        super().__init__()

        self.title("AutoLiv Generant Compaction Station Control")
        self.geometry(f"{1100}x{600}")

        # Grid Layout
        self.grid_columnconfigure((1, 2), weight=0)
        self.grid_rowconfigure((0, 1, 2, 3), weight=0)

        # Radiobuttons
        self.radiobutton_frame = customtkinter.CTkFrame(self)
        self.radiobutton_frame.grid(row=0, column=0, rowspan=1, columnspan=1, padx=(10, 0), pady=(20, 0), sticky="nsew")
        self.radio_var = tkinter.IntVar(value=0)
        self.label_radio_group = customtkinter.CTkLabel(master=self.radiobutton_frame, text="Cam Amplitude", font=('CTkFont', 15))
        self.label_radio_group.grid(row=0, column=1, columnspan=2, padx=10, pady=(20, 20), sticky="ew")
        self.amp_1 = customtkinter.CTkRadioButton(master=self.radiobutton_frame, text="1mm",
                                                  variable=self.radio_var, value=0, command=self.setBounds)
        self.amp_1.grid(row=1, column=0, padx=20, pady=10, sticky="n")
        self.amp_06 = customtkinter.CTkRadioButton(master=self.radiobutton_frame, text="0.6mm",
                                                   variable=self.radio_var, value=1, command=self.setBounds)
        self.amp_06.grid(row=1, column=1, padx=20, pady=10, sticky="n")
        self.amp_04 = customtkinter.CTkRadioButton(master=self.radiobutton_frame, text="0.4mm",
                                                   variable=self.radio_var, value=2, command=self.setBounds)
        self.amp_04.grid(row=1, column=2, padx=20, pady=10, sticky="n")
        self.amp_custom = customtkinter.CTkRadioButton(master=self.radiobutton_frame, text="Custom",
                                                   variable=self.radio_var, value=3, command=self.setBounds)
        self.amp_custom.grid(row=1, column=3, padx=20, pady=10, sticky="n")

        # Frequency Slider
        self.slide_frame = customtkinter.CTkFrame(self)#, fg_color="transparent")
        self.slide_frame.grid(row=1, column=0, rowspan=4, padx=(10, 0), pady=(20, 20), sticky="nsew")
        self.slide_frame.grid_columnconfigure(1, weight=1)
        self.slide_frame.grid_rowconfigure(2, weight=1)
        self.subsysLabel = customtkinter.CTkLabel(self.slide_frame, text="Subsystems Control", font=('CTkFont', 15))
        self.subsysLabel.grid(row=0, column=1, pady=(10,0))
        self.freq_var = "disabled"
        self.freqSlider = customtkinter.CTkSlider(self.slide_frame, orientation="horizontal", state="disabled",
                                                  number_of_steps=250, from_=freqBounds[0], to=freqBounds[1])
        self.freqSlider.bind("<ButtonRelease-1>", self.checkFreqSlider)
        self.freqSlider.grid(row=1, column=1, columnspan=2, padx=(10, 10), pady=(20, 30), sticky="ew")

        self.vibButton = customtkinter.CTkCheckBox(self.slide_frame, text="Enable\nVibration",
                                                   command=self.freqToggle, onvalue=1, offvalue=0)
        self.vibButton.grid(row=1, column=0, padx=(10, 10), pady=(20, 30))

        self.freqEntry = customtkinter.CTkEntry(self.slide_frame, placeholder_text="Hz", state="disabled",
                                                validate='focusout', validatecommand=self.checkFreqEntry, width=50)
        self.freqEntry.grid(row=1, column=3, padx=(10, 10), pady=(20, 30), sticky='w')

        self.freqLabel = customtkinter.CTkLabel(self.slide_frame, text="Hz   ")
        self.freqLabel.grid(row=1, column=4, padx=(0, 10), pady=(20, 30), sticky='w')

        # Force Slider
        self.force_var = "disabled"
        self.forceSlider = customtkinter.CTkSlider(self.slide_frame, orientation="horizontal", state="disabled",
                                                   number_of_steps=250, from_=forceBounds[0], to=forceBounds[1])
        self.forceSlider.bind("<ButtonRelease-1>", self.checkForceSlider)
        self.forceSlider.grid(row=2, column=1, columnspan=2, padx=(10, 10), pady=(30, 30), sticky="ew")

        self.compactButton = customtkinter.CTkCheckBox(self.slide_frame, text="Enable\nCompaction",
                                                       command=self.forceToggle, onvalue=1, offvalue=0)
        self.compactButton.grid(row=2, column=0, padx=(10, 10), pady=(30, 30))

        self.forceEntry = customtkinter.CTkEntry(self.slide_frame, placeholder_text="N", state="disabled",
                                                 validate='focusout', validatecommand=self.checkForceEntry, width=50)
        self.forceEntry.grid(row=2, column=3, padx=(10, 10), pady=(30, 30), sticky='w')

        self.forceLabel = customtkinter.CTkLabel(self.slide_frame, text="N    ")
        self.forceLabel.grid(row=2, column=4, padx=(0, 10), pady=(30, 30), sticky='e')

        # Rotation Slider
        self.rotate_var = "disabled"
        self.rotateSlider = customtkinter.CTkSlider(self.slide_frame, orientation="horizontal", state="disabled",
                                                    number_of_steps=250, from_=0, to=50)
        self.rotateSlider.bind("<ButtonRelease-1>", self.checkRotateSlider)
        self.rotateSlider.grid(row=3, column=1, columnspan=2, padx=(10, 10), pady=(30, 30), sticky="ew")

        self.rotateButton = customtkinter.CTkCheckBox(self.slide_frame, text="Enable\nRotation",
                                                      command=self.rotateToggle, onvalue=1, offvalue=0)
        self.rotateButton.grid(row=3, column=0, padx=(10, 10), pady=(30, 30))

        self.rotateEntry = customtkinter.CTkEntry(self.slide_frame, placeholder_text="rad/s", state="disabled",
                                                  validate='focusout', validatecommand=self.checkRotateEntry, width=50)
        self.rotateEntry.grid(row=3, column=3, padx=(10, 10), pady=(30, 30), sticky='w')

        self.rotateLabel = customtkinter.CTkLabel(self.slide_frame, text="rad/s")
        self.rotateLabel.grid(row=3, column=4, padx=(0, 10), pady=(30, 30), sticky='e')

        # Run Time
        self.time_var = "disabled"
        self.timeSlider = customtkinter.CTkSlider(self.slide_frame, orientation="horizontal", state="disabled",
                                                  number_of_steps=250, from_=timeBounds[0], to=timeBounds[1])
        self.timeSlider.bind("<ButtonRelease-1>", self.checkTimeSlider)
        self.timeSlider.grid(row=4, column=1, columnspan=2, padx=(10, 10), pady=(30, 20), sticky="ew")

        self.timeButton = customtkinter.CTkCheckBox(self.slide_frame, text="Enable\nRun Time", command=self.timeToggle,
                                                    onvalue=1, offvalue=0)
        self.timeButton.grid(row=4, column=0, padx=(10, 10), pady=(30, 20))

        self.timeEntry = customtkinter.CTkEntry(self.slide_frame, placeholder_text="sec.", state="disabled",
                                                validate='focusout', validatecommand=self.checkTimeEntry, width=50)
        self.timeEntry.grid(row=4, column=3, padx=(10, 10), pady=(30, 20), sticky='w')

        self.timeLabel = customtkinter.CTkLabel(self.slide_frame, text="sec. ")
        self.timeLabel.grid(row=4, column=4, padx=(0, 10), pady=(30, 20), sticky='e')

        # File Name Frame
        self.filename_frame = customtkinter.CTkFrame(self, width=100)
        self.filename_frame.grid(row=0, column=1, rowspan=1, padx=(20, 0), pady=(20, 0), sticky="nsew")
        self.results_var = tkinter.IntVar(value=0)

        self.filenameEntry = customtkinter.CTkEntry(master=self.filename_frame, placeholder_text='File_Name',
                                                    validate='focusout', width=200, height=30,
                                                    validatecommand=self.checkFilename)
        self.filenameEntry.grid(row=1, column=0, padx=10, pady=(5, 20), sticky='nsew', columnspan=2)
        self.filenameLabel = customtkinter.CTkLabel(self.filename_frame, text='File Name', font=('CTkFont', 15))
        self.filenameLabel.grid(row=0, column=0, padx=10, pady=(20, 5), sticky='sw')
        self.saveData_var = tkinter.IntVar(value=0)
        self.fileButtonSave = customtkinter.CTkRadioButton(master=self.filename_frame, text="Save Data",
                                                           variable=self.saveData_var, value=0, command=self.saveFile)
        self.fileButtonSave.grid(row=2, column=0, padx=15, pady=(10, 20), sticky='nw')
        self.fileButtonNoSave = customtkinter.CTkRadioButton(master=self.filename_frame, text="Run Without Saving",
                                                             variable=self.saveData_var, value=1, command=self.saveFile)
        self.fileButtonNoSave.grid(row=2, column=1, padx=10, pady=(10, 20), sticky='nw')

        #Program Info Frame
        self.info_frame = customtkinter.CTkFrame(self)
        self.info_frame.grid(row=1, column=1, padx=(20, 0), pady=(20, 0), sticky="nsew")


        #Print Program Info Using Text Box
        programInfo = '1) Select Cam amplitude (in mm)\n' \
                      '     *This will automatically set bounds for frequencies .....\n\n' \
                      "2) Select which subsystems to enable using checkboxes\n" \
                      "     Vibration: Enables motor to run at desired frequency (in Hz),\n" \
                      "                        inducing vibration of the module\n" \
                      "     Compaction: Enables pneumatic piston to apply vertical \n" \
                      "                              compression force (in N) to the generant\n" \
                      "     Rotation: Enables servo motor to rotate module at ___rad/s in a \n" \
                      "                      fixed direction\n" \
                      "                      *Adjusting the rotation frequency (in Hz) will change the \n" \
                      "                      rotation direction at the desired rate\n" \
                      "     Run Time: Enables the user to select a desired run time (in sec.)\n" \
                      "                         *With 'Run Time' disabled, the system will run until the \n" \
                      "                          'STOP' button is selected\n\n" \
                      "3) Select if the user would like to save the data\n" \
                      "      *If data is to be saved, enter a filename and measured results \n" \
                      "        will be saved as a .xlsx file\n" \
                      "      *If data is not to be saved, measured results will be displayed \n" \
                      "        without saving and filename may be ignored\n\n" \
                      "4) Once desired variables are entered, select the 'RUN' button to run \n" \
                      "      the system. \n" \
                      "      *Once the desired time is reached,\n" \
                      "        or the 'STOP' button is selected, the measured results will be \n" \
                      "        displayed and the data will be saved, if enabled\n"

        self.infoFrameText = customtkinter.CTkTextbox(master=self.info_frame, height=250)
        self.infoFrameText.pack(fill='both', padx=10, pady=5)
        self.infoFrameText.insert(0.0, programInfo)
        self.infoFrameText.configure(state='disabled')

        #Run/Stop Button
        self.button_frame = customtkinter.CTkFrame(self, fg_color='transparent')
        self.button_frame.grid(row=3, column=1, columnspan=1, padx=10, pady=10)
        self.runButton = customtkinter.CTkButton(master=self.button_frame, width=225, height=75, corner_radius=25, fg_color='green',
                                                 border_color='#006400', hover_color='#228B22', border_width=5, font=('CTkFont', 20),
                                                 text='RUN', command=self.runButtonFunc)
        self.runButton.grid(row=3, column=1, padx=(25, 10), pady=0, sticky='sw')

        self.stopButton = customtkinter.CTkButton(master=self.button_frame, width=225, height=75, corner_radius=25, fg_color='#EE2C2C',
                                                  border_color='#B22222', hover_color='#FF3030', border_width=5, font=('CTkFont', 20),
                                                  text='STOP', command=self.stopButtonFunc, state='disabled')
        self.stopButton.grid(row=3, column=2, padx=10, pady=0, sticky='se')

        self.resultsWindow = None
    # Functions to set bounds due to amplitude
    def setBounds(self):
        global freqBounds, forceBounds, rotSpeedBounds
        radio_var = self.radio_var.get()
        self.freqEntry.delete(0, len(self.freqEntry.get()))
        self.forceEntry.delete(0, len(self.forceEntry.get()))
        self.rotateEntry.delete(0, len(self.rotateEntry.get()))
        self.timeEntry.delete(0, len(self.timeEntry.get()))
        if radio_var == 0:
            amp = '1mm'
            freqBounds = [freqBounds1[0], freqBounds1[1]]
            forceBounds = [forceBounds1[0], forceBounds1[1]]
            rotSpeedBounds = [rotSpeedBounds1[0], rotSpeedBounds1[1]]
        elif radio_var == 1:
            amp = '0.6mm'
            freqBounds = [freqBounds06[0], freqBounds06[1]]
            forceBounds = [forceBounds06[0], forceBounds06[1]]
            rotSpeedBounds = [rotSpeedBounds06[0], rotSpeedBounds06[1]]
        elif radio_var == 2:
            amp = '0.4mm'
            freqBounds = [freqBounds04[0], freqBounds04[1]]
            forceBounds = [forceBounds04[0], forceBounds04[1]]
            rotSpeedBounds = [rotSpeedBounds04[0], rotSpeedBounds04[1]]
        elif radio_var == 3:
            amp = 'custom'
            freqBounds = [freqBoundsCustom[0], freqBoundsCustom[1]]
            forceBounds = [forceBoundsCustom[0], forceBoundsCustom[1]]
            rotSpeedBounds = [rotSpeedBoundsCustom[0], rotSpeedBoundsCustom[1]]
            print("Custom cam amplitude selected\n*Make changes to allow for amplitude entry*")
        else:
            print('Something wrong with RadioButton function setBounds')
        self.freqSlider.configure(number_of_steps=250, from_=freqBounds[0], to=freqBounds[1])
        self.forceSlider.configure(number_of_steps=250, from_=forceBounds[0], to=forceBounds[1])
        self.rotateSlider.configure(number_of_steps=250, from_=rotSpeedBounds[0], to=rotSpeedBounds[1])
        print('Subsystem bounds for ' + amp + ' amplitude')
        print('freqBounds:[' + str(freqBounds[0]) + ', ' + str(freqBounds[1]) + ']')
        print('forceBounds:[' + str(forceBounds[0]) + ', ' + str(forceBounds[1]) + ']')
        print('rotSpeedBounds:[' + str(rotSpeedBounds[0]) + ', ' + str(rotSpeedBounds[1]) + ']\n')

    def saveFile(self):
        fileNameValid = self.checkFilename()
        if self.saveData_var.get() == 0:
            if fileNameValid:
                fileName = self.filenameEntry.get()
                wb = Workbook()
                wb.save(str(fileName)+'.xlsx')
                sheet = wb.active
                #Initialize sheet labels
                sheet.row_dimensions[2].height = 27
                sheet.column_dimensions['C'].width = 11
                sheet.column_dimensions['E'].width = 10

                sheet['D2'].value = 'Settings'
                sheet['E2'] = 'Max\nMeasured'
                sheet['E2'].alignment = Alignment(wrapText=True)
                sheet['C3'].value = 'Freq.'
                sheet['C4'].value = 'Force'
                sheet['C5'].value = 'Rot. Freq'
                sheet['C6'].value = 'Time'
                sheet['F3'].value = 'Hz'
                sheet['F4'].value = 'N'
                sheet['F5'].value = 'Hz'
                sheet['F6'].value = 'Sec'
                sheet['C8'].value = 'Freq.'
                sheet['C9'].value = 'Force'
                sheet['C10'].value = 'Rot. Speed'
                sheet['C11'].value = 'Time'

                #Insert values into sheet
                sheet['D3'].value = float(self.freqEntry.get())
                sheet['D4'].value = float(self.forceEntry.get())
                sheet['D5'].value = float(self.rotateEntry.get())
                sheet['D6'].value = float(self.timeEntry.get())
                sheet['E3'].value = maxFreq
                sheet['E4'].value = maxForce
                sheet['E5'].value = maxRotation
                sheet['E6'].value = maxTime

                for i in range(len(freqMeasures)):
                    sheet.cell(row=8, column=i+4).value = freqMeasures[i]

                for i in range(len(forceMeasures)):
                    sheet.cell(row=9, column=i+4).value = forceMeasures[i]

                for i in range(len(rotationMeasures)):
                    sheet.cell(row=10, column=i+4).value = rotationMeasures[i]

                for i in range(len(t)):
                    sheet.cell(row=11, column=i+4).value = t[i]
                wb.save(str(fileName)+'.xlsx')
            else:
                print('Invalid filename')
        elif self.saveData_var == 1:
            print("Not Saving Data")

        else:
            print('Error with save data radio buttons')


    # Functions to disable sliders and entry boxes when related function is not enabled by the user

    def freqToggle(self):

        if self.vibButton.get() == 1:
            self.freqSlider.configure(state="normal")
            self.freqEntry.configure(state="normal")
        else:
            self.freqSlider.configure(state="disabled")
            self.freqEntry.delete(0, len(self.freqEntry.get()))
            self.freqEntry.configure(state="disabled")

    def forceToggle(self):
        if self.compactButton.get() == 1:
            self.forceSlider.configure(state="normal")
            self.forceEntry.configure(state="normal")
        else:
            self.forceSlider.configure(state="disabled")
            self.forceEntry.delete(0, len(self.forceEntry.get()))
            self.forceEntry.configure(state="disabled")

    def rotateToggle(self):
        if self.rotateButton.get() == 1:
            self.rotateSlider.configure(state="normal")
            self.rotateEntry.configure(state="normal")
        else:
            self.rotateSlider.configure(state="disabled")
            self.rotateEntry.delete(0, len(self.rotateEntry.get()))
            self.rotateEntry.configure(state="disabled")

    def timeToggle(self):
        if self.timeButton.get() == 1:
            self.timeSlider.configure(state="normal")
            self.timeEntry.configure(state="normal")
        else:
            self.timeSlider.configure(state="disabled")
            self.timeEntry.delete(0, len(self.timeEntry.get()))
            self.timeEntry.configure(state="disable")

    # Functions to sync slider and entry values

    def checkFreqEntry(self):
        try:                                            #Check if entry is numeric (int or float)
            val = int(self.freqEntry.get())
            val1 = True
        except ValueError:
            try:
                val = float(self.freqEntry.get())
                val1 = True
            except ValueError:
                val1 = False

        if val1 == True:                                    #If entry is numeric, set slider to entry value
            if freqBounds[0] <= val <= freqBounds[1]:
                self.freqSlider.set(val)
                return True
        else:                                                       #If entry is not numeric or out of freqBounds, print an error with the proper bounds
            self.freqEntry.delete(0, len(self.freqEntry.get()))
            print('Invalid frequency')
            print('Enter a value between ' + str(freqBounds[0]) + ' and ' + str(freqBounds[1]) + 'Hz')
            return False

    def checkFreqSlider(self, second):                              #When the freqSlider is changed, update the freqEntry box to match
        if isinstance(self.freqEntry.get(), int):
            if self.freqSlider.get() != int(self.freqEntry.get()):
                self.freqEntry.delete(0, len(self.freqEntry.get()))
                self.freqEntry.insert(self.freqSlider.get)
                return True
        else:
            self.freqEntry.delete(0, len(self.freqEntry.get()))
            self.freqEntry.insert(0, round(self.freqSlider.get(), 1))
            return True

    def checkForceEntry(self):
        try:                                            #Check if entry is numeric (int or float)
            val = int(self.forceEntry.get())
            val1 = True
        except ValueError:
            try:
                val = float(self.forceEntry.get())
                val1 = True
            except ValueError:
                val1 = False

        if val1 == True:                                    #If entry is numeric, set slider to entry value
            if forceBounds[0] <= val <= forceBounds[1]:
                self.forceSlider.set(val)
                return True
        else:                                                       #If entry is not numeric or out of freqBounds, print an error with the proper bounds
            self.forceEntry.delete(0, len(self.forceEntry.get()))
            print('Invalid compaction force')
            print('Enter a value between ' + str(forceBounds[0]) + ' and ' + str(forceBounds[1]) + 'N')
            return False

    def checkForceSlider(self, second):                              #When the forceSlider is changed, update the forceEntry box to match
        if isinstance(self.forceEntry.get(), int):
            if self.forceSlider.get() != int(self.forceEntry.get()):
                self.forceEntry.delete(0, len(self.forceEntry.get()))
                self.forceEntry.insert(self.forceSlider.get)
                return True
        else:
            self.forceEntry.delete(0, len(self.forceEntry.get()))
            self.forceEntry.insert(0, round(self.forceSlider.get(), 1))
            return True

    def checkRotateEntry(self):
        try:                                            #Check if entry is numeric (int or float)
            val = int(self.rotateEntry.get())
            val1 = True
        except ValueError:
            try:
                val = float(self.rotateEntry.get())
                val1 = True
            except ValueError:
                val1 = False

        if val1 == True:                                    #If entry is numeric, set slider to entry value
            if rotSpeedBounds[0] <= val <= rotSpeedBounds[1]:
                self.rotateSlider.set(val)
                return True
        else:                                                       #If entry is not numeric or out of freqBounds, print an error with the proper bounds
            self.rotateEntry.delete(0, len(self.rotateEntry.get()))
            print('Invalid rotation speed')
            print('Enter a value between ' + str(rotSpeedBounds[0]) + ' and ' + str(rotSpeedBounds[1]) + 'rad/s')
            return False

    def checkRotateSlider(self, second):                              #When the freqSlider is changed, update the freqEntry box to match
        if isinstance(self.rotateEntry.get(), int):
            if self.rotateSlider.get() != int(self.rotateEntry.get()):
                self.rotateEntry.delete(0, len(self.rotateEntry.get()))
                self.rotateEntry.insert(self.rotateSlider.get)
                return True
        else:
            self.rotateEntry.delete(0, len(self.rotateEntry.get()))
            self.rotateEntry.insert(0, round(self.rotateSlider.get(), 1))
            return True

    def checkTimeEntry(self):
        try:                                            #Check if entry is numeric (int or float)
            val = int(self.timeEntry.get())
            val1 = True
        except ValueError:
            try:
                val = float(self.timeEntry.get())
                val1 = True
            except ValueError:
                val1 = False

        if val1 == True:                                    #If entry is numeric, set slider to entry value
            if timeBounds[0] <= val <= timeBounds[1]:   #slider from_ to
                self.timeSlider.set(val)    #set slider val
                return True
            elif val >= 0:
                self.timeSlider.configure(to=val)   #change slider from_ to
                self.timeSlider.set(val)#set slider value to entry
                return True
        else:                                                       #If entry is not numeric or out of freqBounds, print an error with the proper bounds
            self.freqEntry.delete(0, len(self.freqEntry.get()))
            print('Invalid frequency')
            print('Enter a value between ' + str(freqBounds[0]) + ' and ' + str(freqBounds[1]) + 'Hz')
            return False

    def checkTimeSlider(self, second):                              #When the freqSlider is changed, update the freqEntry box to match
        if isinstance(self.timeEntry.get(), int):
            if self.timeSlider.get() != int(self.timeEntry.get()):
                self.timeEntry.delete(0, len(self.timeEntry.get()))
                self.timeEntry.insert(self.timeSlider.get)
                return True
        else:
            self.timeEntry.delete(0, len(self.timeEntry.get()))
            self.timeEntry.insert(0, round(self.timeSlider.get(), 1))
            return True

    def checkFilename(self):
        filename = self.filenameEntry.get()
        if not re.findall(r'[^A-Za-z0-9_\-\\]', filename):
            if re.findall(r'[^0-9_\-\\]', filename[0]):
                fileNameValid = True
            else:
                fileNameValid = False
        else:
            fileNameValid = False
        return fileNameValid

    def runButtonFunc(self):
        #Start all subsystems
        self.runButton.configure(state='disabled')
        self.stopButton.configure(state='normal', hover=True)
        self.sendToClearcore()

        global clearcoreResults

        while clearcore.inWaiting():
            #Add check for stop button press
            clearcoreResults = clearcore.read(clearcore.inWaiting()).decode()
            self.stopButtonFunc()
        return

    def stopButtonFunc(self):
        #Stop all subsystems
        self.stopButton.configure(state='disabled')
        self.runButton.configure(state='normal', hover=True)
        self.openResultsWindow()
        self.saveFile()
        return

    def openResultsWindow(self):
        if self.resultsWindow is None or not self.resultsWindow.winfo_exists():
            self.resultsWindow = ProcessResults(self)
            self.resultsWindow.focus()

        else:
            self.resultsWindow.focus()

    def sendToClearcore(self):

        if self.vibButton.get() == 1:
            setFreq = self.freqEntry.get()
        else:
            setFreq = 0
        if self.compactButton.get() == 1:
            setForce = self.forceEntry.get()
        else:
            setForce = 0
        if self.rotateButton.get() == 1:
            setRotation = self.rotateEntry.get()
        else:
            setRotation = 0
        if self.timeButton.get() == 1:
            setTime = self.timeEntry.get()
        else:
            setTime = 0
        sendComm = 'Freq:' + str(setFreq) + ',Force:' + str(setForce) + ',Rotation:' + str(setRotation) + \
                   ',Time:' + str(setTime)
        clearcore.write(str.encode(sendComm))


if __name__ == "__main__":
    app = App()
    app.mainloop()